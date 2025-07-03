from openpyxl import load_workbook
import pandas as pd

# 엑셀 파일 경로
excel_file = r"D:\workspace\project\KG_AI_Project\resource\raw_files\기본통계_소장및구독자료.xlsx"
wb = load_workbook(excel_file, data_only=True)
sheet_names = wb.sheetnames

for sheet in sheet_names:
    ws = wb[sheet]
    year = int(sheet)
    
    max_col = ws.max_column
    max_row = ws.max_row

    # ✅ A~F열: 고정 컬럼명
    fixed_names = ['번호','학교명','학교유형','설립','지역','대학규모']
    column_names = fixed_names.copy()

    # ✅ G열부터 병합 셀 기반 컬럼 자동 생성
    for col in range(7, max_col + 1):
        parts = []
        row = 5
        top_cell = ws.cell(row=row, column=col).value
        if top_cell == "":
            break  # 공란 셀이면 컬럼 생성 중지

        while row <= max_row:
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                break  # 숫자 만나면 컬럼 종료
            elif val is None or (isinstance(val, str) and val.strip() == ""):
                row += 1
                continue
            elif isinstance(val, str):
                parts.append(val.strip())
            row += 1

        name = "-".join(parts) if parts else f"열{col}"
        column_names.append(name)

    # ✅ 데이터 시작행 찾기 (A열 기준)
    data_start = None
    for row in range(6, max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, (int, float)):
            data_start = row
            break

    if not data_start:
        print(f"⚠️ {sheet} 시트에는 데이터 시작 행을 찾을 수 없습니다. 건너뜁니다.")
        continue

    # ✅ 실제 데이터 열 개수 찾기 (G열 기준)
    actual_data_col_count = 6  # A~F열 포함
    for col in range(7, max_col + 1):
        val = ws.cell(row=data_start, column=col).value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            break
        actual_data_col_count += 1

    # 컬럼명 정리
    column_names = column_names[:actual_data_col_count]

    # ✅ 데이터 종료행 찾기
    data_end = data_start
    while data_end <= max_row:
        val = ws.cell(row=data_end, column=1).value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            break
        data_end += 1

    # ✅ 데이터 수집
    records = []
    for row in range(data_start, data_end):
        record = [year]  # 연도 컬럼
        for col in range(1, actual_data_col_count + 1):
            record.append(ws.cell(row=row, column=col).value)
        records.append(record)

    # ✅ DataFrame 생성 및 저장
    df = pd.DataFrame(records, columns=['연도'] + column_names)
    filename = f"Num01_소장및구독자료_{year}.csv"
    df.to_csv(filename, index=False, encoding='utf-8-sig')
    print(f"✅ {filename} 저장 완료! (컬럼 {len(column_names)+1}개, 데이터 {len(records)}행)")