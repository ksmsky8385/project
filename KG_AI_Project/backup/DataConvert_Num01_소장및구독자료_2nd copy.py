import pandas as pd
from openpyxl import load_workbook

# 기본통계_소장및구독자료.xlsx 데이터시트 csv 파일 변환 스크립트 두번째.
# 2024년도는 추가 변환 후 저장 수정코드.

# 엑셀 파일 경로
excel_file = r'D:\workspace\project\KG_AI_Project\resource\raw_files\기본통계_소장및구독자료.xlsx'

# 시트 이름 목록 가져오기
wb = load_workbook(excel_file, read_only=True)
sheet_names = wb.sheetnames

# 기본 컬럼명 (2023년도까지)
columns_default = [
    '연도','번호','학교명','학교유형','설립','지역','대학규모',
    '도서자료-종수-국내서','도서자료-종수-국외서','도서자료-종수-계',
    '도서자료-책수-국내서','도서자료-책수-국외서','도서자료-책수-계',
    '전자자료-전자저널-패키지','전자자료-웹DB-패키지','전자자료-eBook-종수',
    '전자자료-동영상강의자료-패키지','전자자료-합계-패키지','전자자료-합계-종수',
    '연속간행물-국내','연속간행물-국외','연속간행물-합계',
    '연간 장서 증가 책수-구입','연간 장서 증가 책수-기증','연간 장서 증가 책수-합계',
    '연간 장서 증가 종수','연간 장서 폐기 책수',
    '소장도서수','재학생수','재학생 1인당 소장도서수',
    '연간장서증가책수','소장도서수','연간장서증가율'
]

# 2024년도 컬럼명 (합계-종수 없음, eBook 항목명 변경)
columns_2024 = [
    '연도','번호','학교명','학교유형','설립','지역','대학규모',
    '도서자료-종수-국내서','도서자료-종수-국외서','도서자료-종수-계',
    '도서자료-책수-국내서','도서자료-책수-국외서','도서자료-책수-계',
    '전자자료-전자저널-패키지','전자자료-웹DB-패키지','전자자료-eBook-패키지',
    '전자자료-동영상강의자료-패키지','전자자료-합계-패키지',
    '연속간행물-국내','연속간행물-국외','연속간행물-합계',
    '연간 장서 증가 책수-구입','연간 장서 증가 책수-기증','연간 장서 증가 책수-합계',
    '연간 장서 증가 종수','연간 장서 폐기 책수',
    '소장도서수','재학생수','재학생 1인당 소장도서수',
    '연간장서증가책수','소장도서수','연간장서증가율'
]

# 변환 대상 컬럼
float_columns = ['재학생 1인당 소장도서수', '연간장서증가율']

# 시트별 처리
for sheet in sheet_names:
    df = pd.read_excel(excel_file, sheet_name=sheet, header=None, engine='openpyxl')

    # 2024년도는 열 수가 다르므로 분기 처리
    if sheet == '2024':
        subset = df.iloc[7:468, 0:31]  # 2024: 총 32열 (0~30)
        subset.insert(0, '연도', int(sheet))
        subset.columns = columns_2024
    else:
        subset = df.iloc[7:468, 0:32]  # 기존: 총 33열 (0~31)
        subset.insert(0, '연도', int(sheet))
        subset.columns = columns_default

    # float 형식 지정
    for col in float_columns:
        if col in subset.columns:
            subset[col] = pd.to_numeric(subset[col], errors='coerce').astype(float)

    # 저장
    filename = f'Num01_소장및구독자료_{sheet}.csv'
    subset.to_csv(filename, index=False, encoding='utf-8-sig')

print("✅ 모든 연도 저장 완료! (2024년도 별도 컬럼명 적용)")