import pandas as pd
from openpyxl import load_workbook

# 기본통계_예산및결산.xlsx 데이터시트 csv 파일 변환 스크립트 첫번째.
# 최근 3년간 자료구입비 컬럼명 개별파일 간 지정 반영.

# 엑셀 경로
excel_file = r'D:\workspace\project\KG_AI_Project\resource\raw_files\기본통계_예산및결산.xlsx'

# 시트 이름 가져오기
wb = load_workbook(excel_file, read_only=True)
sheet_names = wb.sheetnames

# 고정 컬럼명 (앞쪽 30개)
fixed_columns = [
    '번호','학교명','학교유형','설립','지역','대학규모',
    '결산-대학총결산','결산-도서자료 구입비','결산-연속간행물 구입비','결산-비도서자료 구입비',
    '결산-전자자료 구입비-전자저널','결산-전자자료 구입비-웹DB','결산-전자자료 구입비-ebook','결산-전자자료 구입비-기타 전자자료',
    '결산-자료구입비계',
    '예산-대학총예산','예산-도서자료 구입비','예산-연속간행물 구입비','예산-비도서자료 구입비',
    '예산-전자자료 구입비-전자저널','예산-전자자료 구입비-웹DB','예산-전자자료 구입비-ebook','예산-전자자료 구입비-기타 전자자료',
    '예산-자료구입비계',
    '결산-자료구입비','재학생수','재학생 1인당 자료구입비','결산-자료구입비','결산-대학총결산',
    '대학총결산 대비 자료구입비 비율'
]

# 실수형 컬럼
float_columns = [
    '재학생 1인당 자료구입비',
    '대학총결산 대비 자료구입비 비율',
    '최근 3년간 자료구입비 증가율(결산)'
]

# 시트별 처리
for sheet in sheet_names:
    df = pd.read_excel(excel_file, sheet_name=sheet, header=None, engine='openpyxl')
    subset = df.iloc[7:, 0:len(fixed_columns) + 3]  # 총 33열 추출

    # 먼저 연도 숫자 추출
    year = int(sheet)

    # ✨ 동적 컬럼명 생성
    a_col = f'자료구입비(결산)-{year - 2}년(A)'
    b_col = f'자료구입비(결산)-{year - 1}년(B)'
    c_col = f'자료구입비(결산)-{year}년(C)'
    growth_col = '최근 3년간 자료구입비 증가율(결산)'

    # 전체 컬럼명 (연도 추가 포함해서 총 35개)
    final_columns = ['연도'] + fixed_columns + [a_col, b_col, c_col, growth_col]

    # ✏️ subset 슬라이싱 시 컬럼 수 정확히 맞추기
    subset = df.iloc[7:, 0:len(final_columns) - 1]  # 34열 슬라이스 → 연도 열 삽입하면 35열 맞춤

    # ✨ 연도 열 삽입
    subset.insert(0, '연도', year)

    # 🔧 컬럼명 지정
    subset.columns = final_columns

    # float 변환
    for col in float_columns + [growth_col]:
        subset[col] = pd.to_numeric(subset[col], errors='coerce').astype(float)

    # 저장
    filename = f'Num03_예산및결산_{year}.csv'
    subset.to_csv(filename, index=False, encoding='utf-8-sig')

print("✅ 예산/결산 CSV 변환 완료: 연도별 3년치 컬럼명 자동 반영됨!")