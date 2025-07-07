import pandas as pd
from openpyxl import load_workbook

# 기본통계_소장및구독자료.xlsx 데이터시트 csv 파일 변환 스크립트 첫번째.
# 2024년도 컬럼오류로 인한 오류코드

# 엑셀 파일 경로
excel_file = r'D:\workspace\project\KG_AI_Project\resource\raw_files\기본통계_소장및구독자료.xlsx'

# 시트 목록 가져오기
wb = load_workbook(excel_file, read_only=True)
sheet_names = wb.sheetnames

# 고정 컬럼명
columns = [
    '연도','번호','학교명','학교유형','설립','지역','대학규모',
    '도서자료-종수-국내서','도서자료-종수-국외서','도서자료-종수-계',
    '도서자료-책수-국내서','도서자료-책수-국외서','도서자료-책수-계',
    '전자자료-전자저널-패키지','전자자료-웹DB-패키지','전자자료-eBook-종수',
    '전자자료-동영상강의자료-패키지','전자자료-합계-패키지','전자자료-합계-종수',
    '연속간행물-국내','연속간행물-국외','연속간행물-합계',
    '연간 장서 증가 책수-구입','연간 장서 증가 책수-기증','연간 장서 증가 책수-합계',
    '연간 장서 증가 종수','연간 장서 폐기 책수',
    '소장도서수','재학생수','재학생 1인당 소장도서수',
    '연간장서증가책수','소장도서수','연간장서증가율'
]

# 변환 대상 컬럼
float_columns = ['재학생 1인당 소장도서수', '연간장서증가율']

# 시트별 처리
for sheet in sheet_names:
    # 시트 로딩
    df = pd.read_excel(excel_file, sheet_name=sheet, header=None, engine='openpyxl')
    subset = df.iloc[7:468, 0:32]

    # 연도 열 추가
    subset.insert(0, '연도', int(sheet))  # 시트명이 숫자 연도일 경우

    # 컬럼명 지정
    subset.columns = columns

    # float 지정 컬럼만 실수형으로 변환
    for col in float_columns:
        subset[col] = pd.to_numeric(subset[col], errors='coerce').astype(float)

    # CSV 저장
    filename = f'Num01_소장및구독자료_{sheet}.csv'
    subset.to_csv(filename, index=False, encoding='utf-8-sig')

print("✅ AC, AF열만 float형으로 변환되어 CSV 저장 완료!")