import os
import pandas as pd
from openpyxl import load_workbook

class ExcelToCSVConverter_ver1:
    def __init__(self, file_path, output_prefix, save_dir):
        self.file_path = file_path
        self.output_prefix = output_prefix
        self.save_dir = save_dir
        self.fixed_names = ['번호','학교명','학교유형','설립','지역','대학규모']

    def run(self):
        wb = load_workbook(self.file_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            year = int(sheet)
            df = self.extract_dataframe(ws, year)
            if df is not None and not df.empty:
                filename = os.path.join(self.save_dir, f"{self.output_prefix}_{year}.csv")
                df.to_csv(filename, index=False, encoding='utf-8-sig')
                print(f"✅ {(os.path.basename(filename))} 저장 완료. (컬럼 {len(df.columns)}개, 데이터 {len(df)}행)")

        print(f"📂 저장 경로: {self.save_dir}")


    def extract_dataframe(self, ws, year):
        max_col = ws.max_column
        max_row = ws.max_row
        column_names = self.fixed_names.copy()

        # 컬럼명 자동 생성 (G열~)
        for col in range(7, max_col + 1):
            parts = []
            row = 5
            top_cell = ws.cell(row=row, column=col).value
            if top_cell == "":
                break
            while row <= max_row:
                val = ws.cell(row=row, column=col).value
                if isinstance(val, (int, float)):
                    break
                elif val is None or (isinstance(val, str) and val.strip() == ""):
                    row += 1
                    continue
                elif isinstance(val, str):
                    parts.append(val.strip())
                row += 1
            name = "-".join(parts) if parts else f"열{col}"
            column_names.append(name)

        # 데이터 시작행
        data_start = None
        for row in range(6, max_row + 1):
            val = ws.cell(row=row, column=1).value
            if isinstance(val, (int, float)):
                data_start = row
                break
        if not data_start:
            print(f"{year} 시트에는 데이터 시작 행을 찾을 수 없습니다. 건너뜁니다.")
            return None

        # 실제 열 수
        actual_data_col_count = 6
        for col in range(7, max_col + 1):
            val = ws.cell(row=data_start, column=col).value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                break
            actual_data_col_count += 1
        column_names = column_names[:actual_data_col_count]

        # 종료행
        data_end = data_start
        while data_end <= max_row:
            val = ws.cell(row=data_end, column=1).value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                break
            data_end += 1

        # 데이터 수집
        records = []
        for row in range(data_start, data_end):
            record = [year]
            for col in range(1, actual_data_col_count + 1):
                record.append(ws.cell(row=row, column=col).value)
            records.append(record)

        return pd.DataFrame(records, columns=['연도'] + column_names)