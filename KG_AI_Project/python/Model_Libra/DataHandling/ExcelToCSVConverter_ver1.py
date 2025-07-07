import os
import pandas as pd
from openpyxl import load_workbook
from typing import List

class ExcelToCSVConverter_ver1:
    def __init__(self, file_path: str, output_prefix: str, save_dir: str):
        self.file_path = file_path
        self.prefix = output_prefix
        self.save_dir = save_dir
        self.fixed_cols = ["번호", "학교명", "학교유형", "설립", "지역", "대학규모"]
        os.makedirs(save_dir, exist_ok=True)
        self.wb = load_workbook(file_path, data_only=True)

    def detect_data_start_row(self, ws, check_col=1, header_row=5) -> int:
        for r in range(header_row, ws.max_row + 1):
            val = ws.cell(row=r, column=check_col).value
            if isinstance(val, (int, float)):
                return r
        raise ValueError("데이터 시작 행을 찾을 수 없습니다.")

    def build_merged_map(self, ws) -> dict:
        merged_map = {}
        for merged in ws.merged_cells.ranges:
            tl = (merged.min_row, merged.min_col)
            for rr in range(merged.min_row, merged.max_row + 1):
                for cc in range(merged.min_col, merged.max_col + 1):
                    merged_map[(rr, cc)] = tl
        return merged_map

    def extract_dynamic_headers(self, ws, merged_map, data_row, header_start_row=5) -> List[str]:
        max_col = ws.max_column
        dyn = []
        for c in range(len(self.fixed_cols) + 1, max_col + 1):
            parts = []
            seen_keys = set()
            for r in range(header_start_row, data_row):
                coord = (r, c)
                tl = merged_map.get(coord, coord)
                if tl in seen_keys:
                    continue
                seen_keys.add(tl)
                val = ws.cell(row=tl[0], column=tl[1]).value
                if isinstance(val, (int, float)):
                    break
                if val is None or (isinstance(val, str) and not val.strip()):
                    continue
                text = str(val).strip().replace("\n", " ")
                parts.append(text)
            header_name = "_".join(parts)
            dyn.append(header_name)
        return ["연도"] + self.fixed_cols + dyn

    def count_valid_cells_in_row(self, ws, row_idx: int) -> int:
        return sum(1 for cell in ws[row_idx] if cell.value not in (None, ""))

    def run(self):
        print(f"\n저장경로 : {self.save_dir}\n")
        for sheet in self.wb.sheetnames:
            ws = self.wb[sheet]
            start_row = self.detect_data_start_row(ws)
            merged_map = self.build_merged_map(ws)
            cols = self.extract_dynamic_headers(ws, merged_map, start_row, header_start_row=5)
            data_cols_count = self.count_valid_cells_in_row(ws, start_row)

            try:
                year = int(sheet)
            except ValueError:
                print(f"시트명 '{sheet}' → 정수 변환 실패. '연도'에 NULL 입력")
                year = None

            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet,
                header=None,
                skiprows=start_row - 1,
                usecols=range(data_cols_count),
                nrows=ws.max_row - (start_row - 1),
                names=cols[1:1 + data_cols_count],
                dtype=str
            )
            df.insert(0, "연도", year)

            out_csv = os.path.join(self.save_dir, f"{self.prefix}_{sheet}.csv")
            df.to_csv(out_csv, index=False, encoding="utf-8-sig")
            print(f"[{self.prefix}_{sheet}.csv] → 저장 완료. (컬럼 {len(df.columns)}개, 데이터 {len(df)}행)")
