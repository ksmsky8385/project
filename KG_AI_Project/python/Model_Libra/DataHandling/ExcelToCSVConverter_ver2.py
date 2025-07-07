import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

class ExcelToCSVConverter_ver2:
    def __init__(self, file_path, output_prefix, save_dir):
        self.file_path = file_path
        self.output_prefix = output_prefix
        self.save_dir = save_dir
        self.fixed_cols = ["번호", "학교명", "학교유형", "설립", "지역", "대학규모"]
        os.makedirs(save_dir, exist_ok=True)
        self.wb = load_workbook(self.file_path, data_only=True)
        self.ws = self.wb.active

    def build_merged_map(self, ws) -> dict:
        merged_map = {}
        for merged in ws.merged_cells.ranges:
            tl = (merged.min_row, merged.min_col)
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    merged_map[(r, c)] = tl
        return merged_map

    def extract_column_names(self, col_start, col_end, data_start_row) -> list[str]:
        merged_map = self.build_merged_map(self.ws)
        colnames = []
        for c in range(col_start, col_end + 1):
            parts = []
            seen_keys = set()
            for r in range(5, data_start_row):
                coord = (r, c)
                tl = merged_map.get(coord, coord)
                if tl in seen_keys:
                    continue
                seen_keys.add(tl)
                val = self.ws.cell(row=tl[0], column=tl[1]).value
                if isinstance(val, (int, float)):
                    break
                if val is None or (isinstance(val, str) and not val.strip()):
                    continue
                text = str(val).strip().replace("\n", " ")
                parts.append(text)
            header_name = "_".join(parts)
            colnames.append(header_name or f"열{c}")
        return colnames

    def detect_data_start_row(self, check_col=1, header_row=5) -> int:
        for r in range(header_row, self.ws.max_row + 1):
            val = self.ws.cell(row=r, column=check_col).value
            if isinstance(val, (int, float)):
                return r
        return None

    def detect_data_end_row(self, data_start: int, check_col=1) -> int:
        r = data_start
        while r <= self.ws.max_row:
            val = self.ws.cell(row=r, column=check_col).value
            if val is None or (isinstance(val, str) and not val.strip()):
                break
            r += 1
        return r

    def extract_year_slices(self) -> dict:
        from collections import OrderedDict
        year_map = OrderedDict()
        for cell_range in self.ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
            if min_row == 4:
                val = self.ws.cell(row=4, column=min_col).value
                if val and str(val).strip().isdigit():
                    year = int(str(val).strip())
                    year_map[year] = (min_col, max_col)
        return year_map

    def extract_dataframe(self, year: int, col_start: int, col_end: int, data_start: int, data_end: int) -> pd.DataFrame:
        year_columns = self.extract_column_names(col_start, col_end, data_start)
        records = []
        for row in range(data_start, data_end):
            record = [year]
            for col in range(1, 7):
                record.append(self.ws.cell(row=row, column=col).value)
            for col in range(col_start, col_end + 1):
                record.append(self.ws.cell(row=row, column=col).value)
            records.append(record)
        return pd.DataFrame(records, columns=["연도"] + self.fixed_cols + year_columns)

    def run(self):
        print(f"\n저장경로 : {self.save_dir}\n")
        year_map = self.extract_year_slices()
        data_start = self.detect_data_start_row()
        data_end = self.detect_data_end_row(data_start)
        if not data_start:
            print("데이터 시작행을 찾을 수 없습니다.")
            return

        for year in sorted(year_map):
            col_start, col_end = year_map[year]
            df = self.extract_dataframe(year, col_start, col_end, data_start, data_end)
            if df is not None and not df.empty:
                out_csv = os.path.join(self.save_dir, f"{self.output_prefix}_{year}.csv")
                df.to_csv(out_csv, index=False, encoding="utf-8-sig")
                print(f"[{os.path.basename(out_csv)}] →  저장 완료 ({df.shape[0]}행, {df.shape[1]}열)")
