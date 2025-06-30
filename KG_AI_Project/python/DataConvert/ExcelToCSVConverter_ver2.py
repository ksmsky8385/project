import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

class ExcelToCSVConverter_ver2:
    def __init__(self, file_path, output_prefix, save_dir):
        self.file_path = file_path
        self.output_prefix = output_prefix
        self.save_dir = save_dir
        self.fixed_names = ['번호', '학교명', '학교유형', '설립', '지역', '대학규모']

    def run(self):
        os.makedirs(self.save_dir, exist_ok=True)
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb.active

        # 병합 셀 기준 연도별 열 범위 추출
        year_map = {}  # {year: (start_col, end_col)}
        for cell_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))
            if min_row == 4:
                val = ws.cell(row=4, column=min_col).value
                if val and str(val).strip().isdigit():
                    year = int(str(val).strip())
                    year_map[year] = (min_col, max_col)

        # 연도별 CSV 저장
        for year, (col_start, col_end) in year_map.items():
            df = self.extract_dataframe(ws, year, col_start, col_end)
            if df is not None and not df.empty:
                filename = os.path.join(self.save_dir, f"{self.output_prefix}_{year}.csv")
                df.to_csv(filename, index=False, encoding='utf-8-sig')
                print(f"✅ {os.path.basename(filename)} 저장 완료. (컬럼 {len(df.columns)}개, 데이터 {len(df)}행)")
        print(f"📂 저장 경로: {self.save_dir}")

    def extract_dataframe(self, ws, year, col_start, col_end):
        max_row = ws.max_row

        # 연도별 컬럼명 생성
        year_columns = []
        for col in range(col_start, col_end + 1):
            parts = []
            for row in range(5, max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    continue
                if isinstance(val, (int, float)):
                    break
                parts.append(str(val).strip())
            name = "-".join(parts) if parts else f"열{col}"
            year_columns.append(name)

        # 데이터 시작행 찾기
        data_start = None
        for row in range(6, max_row + 1):
            val = ws.cell(row=row, column=1).value
            if isinstance(val, (int, float)):
                data_start = row
                break
        if not data_start:
            print(f"⚠️ {year} 연도 데이터 시작행을 찾지 못했습니다.")
            return None

        # 데이터 종료행
        data_end = data_start
        while data_end <= max_row:
            val = ws.cell(row=data_end, column=1).value
            if val is None or (isinstance(val, str) and val.strip() == ""):
                break
            data_end += 1

        # 데이터 수집
        records = []
        for row in range(data_start, data_end):
            record = [year]
            for col in range(1, 7):  # 고정 정보
                record.append(ws.cell(row=row, column=col).value)
            for col in range(col_start, col_end + 1):  # 해당 연도 데이터
                record.append(ws.cell(row=row, column=col).value)
            records.append(record)

        return pd.DataFrame(records, columns=['연도'] + self.fixed_names + year_columns)